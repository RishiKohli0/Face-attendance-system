import xlsxwriter
from openpyxl import Workbook

name = 0 
ID = 0 
lech = 0 
contact = 0

def ExportProfile (ID,name,lech,contact):
		
		
		
	workbook = xlsxwriter.Workbook(str(ID) + '.xlsx')
	worksheet = workbook.add_worksheet()
		
	expenses = (
		[' ID', ID ],
		['NAME',  name],
		['No. of lech',  lech],
		['contact', contact],
		)
		
		
	row = 0
	col = 0

# Iterate over the data and write it out row by row.
	for item, cost in (expenses):
		worksheet.write(row, col,     item)
		worksheet.write(row, col + 1, cost)
		row += 1

# Write a total using a formula.
	worksheet.write(row, 0, 'Total')
	worksheet.write(row, 1, '=SUM(B1:B4)')

	workbook.close()
def update_worksheet(user):
	
	
	wb = Workbook()
	wb.save(str(user) + '.xlsx')
	# grab the active worksheet
	ws = wb.active
	lech +=1
	ws['A3']
	ws.append[lech]
	wb.save(str(user) + '.xlsx')
def update_content(user,column,row,name,ID,lech,contact):
	wb = Workbook()

	# grab the active worksheet
	ws = wb.active
	row_ID = int (row) 
	row_name = row_ID + 1
	row_lech = row_name + 1
	row_contact = row_lech + 1
	# Data can be assigned directly to cells
	ws[ column + str(row_ID)  ] = ID
	ws[ column + str(row_name)  ] =name 
	ws[ column + str(row_lech)  ] = lech
	ws[ column + str(row_contact)  ] = contact
	# Rows can also be appended
	#ws.append([1, 2, 3])

	# Python types will automatically be converted
	import datetime
	#ws[column + row] = datetime.datetime.now()

	# Save the file
	wb.save(str(user) + '.xlsx')
	
	
		
	

		
#ExportProfile(ID='234',name='sidheshwar',lech=56,contact='+918149032989')
#update_worksheet(ID='234')
#update_content(column='B',row='10',name="pandhare",ID="234",lech="46",contact='+918149032989')
